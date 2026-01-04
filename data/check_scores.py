from openpyxl import load_workbook
import re

file_path = 'SMC Businesses Assessment Tool (1).xlsx'
input_sheet = 'Basic Assessment'

try:
    wb = load_workbook(file_path, data_only=True) # data_only=True to get values of I1..I4
    ws = wb[input_sheet]
    
    # Read Score Reference Table (I1..I5 just in case)
    score_refs = {}
    for r in range(1, 6):
        cell_ref = f"$I${r}"
        val = ws.cell(row=r, column=9).value # I is 9th column
        score_refs[cell_ref] = val
        print(f"{cell_ref} = {val}")

    print("-" * 30)

    # Function to parse formula
    def parse_formula(formula, score_refs):
        if not formula or not isinstance(formula, str):
            return {}
        
        # Matches patterns like: C10="Older than 5 years",$I$1
        # regex: \"([^\"]+)\"\s*,\s*(\$I\$\d+)
        matches = re.findall(r'\"([^\"]+)\"\s*,\s*(\$I\$\d+)', formula)
        
        option_map = {}
        for text, ref in matches:
            score = score_refs.get(ref, 0) # Default 0 or None? User said 15, 10 etc.
            option_map[text] = score
        return option_map

    # Test on a few rows
    wb_formulas = load_workbook(file_path, data_only=False) # Helper to read formulas
    ws_formulas = wb_formulas[input_sheet]
    
    for r in range(10, 15):
        formula = ws_formulas.cell(row=r, column=4).value
        print(f"Row {r} Formula: {formula}")
        parsed = parse_formula(formula, score_refs)
        print(f"Parsed: {parsed}")

except Exception as e:
    print(f"Error: {e}")
