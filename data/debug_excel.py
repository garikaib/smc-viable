from openpyxl import load_workbook
import pandas as pd

file_path = 'SMC Businesses Assessment Tool (1).xlsx'

try:
    wb = load_workbook(file_path)
    ws = wb['Basic Assessment']
    
    print("--- Cell Inspection ---")
    # Row 8 is "Questions" start if 1-based is 7. But let's check rows 7-15 roughly.
    # User said data starts row 7.
    # Let's inspect Column C (index 2) for rows 8, 9, etc. (1-based index in openpyxl)
    
    # Rows 7 to 15
    for r in range(7, 16):
        cell_b = ws.cell(row=r, column=2).value # Question
        cell_c = ws.cell(row=r, column=3).value # Option/Content
        print(f"Row {r}: Q='{cell_b}' | C='{cell_c}'")

    print("\n--- Data Validation Inspection ---")
    for validation in ws.data_validations.dataValidation:
        print(f"Range: {validation.sqref} | Type: {validation.type} | Formula1: {validation.formula1}")

    print("\n--- Searching for 'GREAT' ---")
    found = False
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "GREAT" in cell.value:
                    print(f"Found 'GREAT' in {sheet_name} at {cell.coordinate}: {cell.value}")
                    found = True
    if not found:
        print("String 'GREAT' not found in any cell.")

except Exception as e:
    print(f"Error: {e}")
