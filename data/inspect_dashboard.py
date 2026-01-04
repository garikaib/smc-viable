from openpyxl import load_workbook
import json

file_path = 'SMC Businesses Assessment Tool (1).xlsx'
input_sheet = 'Basic Assessment'

def get_color(cell):
    # Try to get fill color
    if cell.fill and cell.fill.start_color:
        return cell.fill.start_color.rgb # often ARGB
        # sometimes theme or indexed.
    return None

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[input_sheet]
    
    print("--- Dashboard Rows Inspection (40-50) ---")
    for r in range(40, 50):
        row_data = []
        for c in range(1, 5): # Cols A, B, C, D
            cell = ws.cell(row=r, column=c)
            val = cell.value
            color = get_color(cell)
            row_data.append(f"{val} [Color:{color}]")
        print(f"Row {r}: {row_data}")

except Exception as e:
    print(f"Error: {e}")
