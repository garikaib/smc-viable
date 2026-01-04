import pandas as pd
from openpyxl import load_workbook
import json
import re
import time
import datetime
import random

file_path = 'SMC Businesses Assessment Tool (1).xlsx'
input_sheet = 'Basic Assessment'
output_file = 'assessments.json'

def generate_id():
    # Maintain some uniqueness
    time.sleep(0.001) 
    return int(time.time() * 1000)

def main():
    try:
        # Load Data-Only for values
        wb_data = load_workbook(file_path, data_only=True)
        ws_data = wb_data[input_sheet]
        
        # Load Formula version
        wb_formulas = load_workbook(file_path, data_only=False)
        ws_formulas = wb_formulas[input_sheet]

        # 1. Read Score Definitions
        score_refs = {}
        for r in range(1, 5): # I1 to I4
            ref_key = f"$I${r}"
            val = ws_data.cell(row=r, column=9).value 
            score_refs[ref_key] = float(val) if val is not None else 0
        
        questions_list = []

        # Iterate Rows 7 to 36 (Questions only)
        for r in range(7, 37):
            # Read cells
            cell_stage = ws_data.cell(row=r, column=1).value
            cell_question = ws_data.cell(row=r, column=2).value
            cell_content_val = ws_data.cell(row=r, column=3).value
            cell_formula = ws_formulas.cell(row=r, column=4).value
            
            # Skip if no question and no stage (empty row)
            if not cell_question and not cell_stage:
                continue
            
            # Skip clearly non-question header/dashboard rows
            if cell_question == "Your Result:":
                continue
            if isinstance(cell_question, str) and ("Recommendation" in cell_question or "Greater than" in cell_question):
                 continue

            # Extract Guidance (content in brackets)
            guidance = ""
            question_text = str(cell_question) if cell_question else ""
            if "(" in question_text and ")" in question_text:
                start = question_text.rfind("(")
                end = question_text.rfind(")")
                if start != -1 and end != -1 and end > start:
                    guidance = question_text[start+1:end]
            
            # Key Text (Example value or placeholder)
            key_text = ""
            if cell_content_val and "Lorem ipsum" not in str(cell_content_val):
                 key_text = str(cell_content_val)
            
            # Determine Type
            # User rule: "You can tell a question is open ended if column C has lorem ipsum or there are the words type it in in column b."
            is_lorem = isinstance(cell_content_val, str) and "Lorem ipsum" in str(cell_content_val)
            is_type_in = isinstance(cell_question, str) and "Type it in" in str(cell_question)
            
            q_type = "text" # default fallback
            dv = None

            # Check Data Validation
            try:
                cell_ref = f"C{r}"
                for validation in ws_data.data_validations.dataValidation:
                    if cell_ref in validation.sqref:
                        dv = validation
                        break
            except Exception:
                pass

            if is_lorem or is_type_in:
                q_type = "text"
            elif dv and dv.type == 'list':
                q_type = "select"
            else:
                 # Fallback logic?
                 q_type = "text"

            entry = {
                "id": generate_id(),
                "type": q_type,
                "stage": cell_stage if cell_stage else "",
                "indicator": cell_stage if cell_stage else "",
                "text": question_text,
                "options": [],
                "guidance": guidance,
                "key_text": key_text
            }

            if q_type == "select" and dv:
                raw_options = dv.formula1
                if raw_options.startswith('"') and raw_options.endswith('"'):
                    raw_options = raw_options[1:-1]
                
                option_labels = [opt.strip() for opt in raw_options.split(',')]
                
                # Parse Formula for Scores
                formula_scores = {}
                if cell_formula and isinstance(cell_formula, str):
                    matches = re.findall(r'\"([^\"]+)\"\s*,\s*(\$I\$\d+)', cell_formula)
                    for label, ref in matches:
                        formula_scores[label] = score_refs.get(ref)

                for label in option_labels:
                    score = formula_scores.get(label)
                    if score is not None:
                         score = int(score) # Integer per request
                    else:
                         # Ensure nulls don't break strict types if needed, but user didn't specify
                         score = 0 # or None
                    
                    opt_obj = {
                        "label": label,
                        "score": score
                    }
                    entry["options"].append(opt_obj)

            # Extra filtering: User said "rows 1-6 not big deal... start from 7".
            # If stage is empty, it might be an invalid row in the middle?
            # Row 16 was found valid. 
            if not entry["stage"]:
                 # Just to be safe, if no stage, don't include
                 continue

            questions_list.append(entry)

        # Final Structure
        output_json = {
            "version": "1.0.0",
            "exportedAt": datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + "Z",
            "assessments": [
                {
                    "title": "Basic Assessment",
                    "questions": questions_list
                }
            ]
        }
        
        print(json.dumps(output_json, indent=2))
        
        with open(output_file, 'w') as f:
            json.dump(output_json, f, indent=2)
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
