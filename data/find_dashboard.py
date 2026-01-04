from openpyxl import load_workbook

file_path = 'SMC Businesses Assessment Tool (1).xlsx'
input_sheet = 'Basic Assessment'

try:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[input_sheet]
    
    print("--- Searching for Dashboard Keywords ---")
    keywords = ["Greater than", "Brilliant", "Recommendation", "Your Result"]
    
    for r in range(1, 60):
        for c in range(1, 15):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str):
                for k in keywords:
                    if k in val:
                        print(f"Found '{k}' at ({r}, {c}): {val[:50]}...")

except Exception as e:
    print(f"Error: {e}")
