import pandas as pd
from openpyxl import load_workbook
import json
import re

file_path = 'SMC Businesses Assessment Tool (1).xlsx'
input_sheet = 'Basic Assessment'
output_file = 'dashboard.json'

def parse_range(text):
    text = text.lower().strip()
    if "greater than" in text:
        nums = re.findall(r'\d+', text)
        if nums:
            val = int(nums[0])
            return {"operator": "gt", "value": val, "min": val + 1, "max": None}
    elif "less than" in text:
        nums = re.findall(r'\d+', text)
        if nums:
            val = int(nums[0])
            return {"operator": "lt", "value": val, "min": None, "max": val - 1}
    elif "between" in text:
        nums = re.findall(r'\d+', text)
        if len(nums) >= 2:
            val1 = int(nums[0])
            val2 = int(nums[1])
            return {"operator": "between", "min": val1, "max": val2}
    return None

def get_style(index, total):
    # Heuristic based on order: Best -> Worst
    if index == 0:
        return {"color": "green", "variant": "success", "icon": "star"}
    elif index == 1:
        return {"color": "light-green", "variant": "success-light", "icon": "check"}
    elif index == 2:
        return {"color": "orange", "variant": "warning", "icon": "alert"}
    else:
        return {"color": "red", "variant": "danger", "icon": "x"}

def main():
    try:
        # Load Data-Only for values
        wb_data = load_workbook(file_path, data_only=True)
        ws_data = wb_data[input_sheet]

        rules = []
        
        # Rows 31 to 35 approx
        # We need to search for the start row dynamically or hardcode if stable
        # Based on previous find_dashboard, "Recommendation" is at Row 31.
        
        start_row = 31
        
        # Iterate until we hit empty or non-matching
        for r in range(start_row, start_row + 10):
            # Col 2 has Condition, Col 3 has Message
            cond_cell = ws_data.cell(row=r, column=2).value
            msg_cell = ws_data.cell(row=r, column=3).value
            
            if not cond_cell:
                # Maybe offset? previous finding said (31, 2)
                continue
            
            cond_str = str(cond_cell).strip()
            
            # Check if it looks like a rule "Greater than", "Between", "Less than"
            if not any(x in cond_str for x in ["Greater than", "Between", "Less than"]):
                 continue
            
            logic = parse_range(cond_str)
            if not logic:
                continue

            # Check validation of message
            if not msg_cell or len(str(msg_cell)) < 10 or "Between" in str(msg_cell) and len(str(msg_cell)) < 20:
                 continue

            rule = {
                "id": f"rule_{len(rules)+1}",
                "condition_text": cond_str,
                "logic": logic,
                "message": msg_cell,
                "style": get_style(len(rules), 4) # approx total 4
            }
            rules.append(rule)

        dashboard_json = {
            "version": "1.0.0",
            "dashboard_config": {
                "title": "Business Viability Assessment Results",
                "rules": rules
            }
        }

        print(json.dumps(dashboard_json, indent=2))
        
        with open(output_file, 'w') as f:
            json.dump(dashboard_json, f, indent=2)

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
