from openpyxl import load_workbook

file_path = 'SMC Businesses Assessment Tool (1).xlsx'

try:
    wb = load_workbook(file_path)
    ws = wb['Basic Assessment']
    
    print("--- Formula Inspection ---")
    # Check D8, D9...
    for r in range(8, 12):
        cell_d = ws.cell(row=r, column=4)
        print(f"Row {r} Column D Formula: {cell_d.value}")

    print("\n--- Hidden Columns Inspection (Rows 7-10, Cols G-R) ---")
    # Columns G (7) to R (18)
    # Print header if exists (Row 3, 4, 5, 6) and data
    
    # Headers in Row 3 or 4?
    headers = []
    for c in range(7, 19):
         val = ws.cell(row=3, column=c).value
         headers.append(str(val))
    print(f"Row 3 Headers (G-R): {headers}")
    
    headers = []
    for c in range(7, 19):
         val = ws.cell(row=5, column=c).value
         headers.append(str(val))
    print(f"Row 5 Headers (G-R): {headers}")

    # Data
    for r in range(7, 12):
        row_vals = []
        for c in range(7, 19):
            val = ws.cell(row=r, column=c).value
            row_vals.append(str(val))
        print(f"Row {r} (G-R): {row_vals}")

except Exception as e:
    print(f"Error: {e}")
